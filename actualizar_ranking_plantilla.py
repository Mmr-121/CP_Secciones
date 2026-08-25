"""
actualizar_ranking_plantilla.py

Vuelca el informe generado en Ranking_CP conservando el mapeo completo de columnas,
incluyendo Evolución Bruta y Evolución Neta.

A diferencia de versiones anteriores, esta NO borra y reescribe la hoja a ciegas:
antes de tocar nada, lee lo que ya hay en Ranking_CP y guarda, por Código Postal,
las columnas que rellena otra persona a mano (Habitantes, % Unifamiliares,
Competencia, IPJ, Prioridad SEO, Prioridad Ads). Después reconstruye la hoja con
los datos nuevos del pipeline y vuelve a colocar esos valores manuales en la fila
que le corresponda a cada CP, aunque el ranking haya cambiado de orden.

La columna Observaciones es mixta: la parte automática (aviso de baja cobertura
de renta) la genera el script; cualquier texto adicional que alguien haya añadido
a mano se conserva y se re-adjunta al aviso automático si sigue aplicando.
"""

import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent

INFORME = BASE_DIR / "Informe_Renta_Urbanizacion_por_CP.xlsx"
PLANTILLA = BASE_DIR / "Plantilla_Profesional_IPJ_Alicante.xlsx"
HOJA = "Ranking_CP"

COLUMNAS = [
    "Ranking", "Código Postal", "Municipio", "Codigo INE",
    "Renta Bruta media (€)", "Evolución bruta", "Renta Neta media (€)",
    "Evolucion Neta", "Habitantes", "% Unifamiliares", "Urbanizaciones",
    "Competencia", "IPJ", "Prioridad SEO", "Prioridad Ads", "Observaciones",
]
COL_IDX = {nombre: i + 1 for i, nombre in enumerate(COLUMNAS)}

# Columnas que rellena otra persona / otro proceso a mano y que este script
# nunca debe pisar sin conservar antes lo que hubiera.
COLUMNAS_MANUALES = [
    "Habitantes", "% Unifamiliares", "Competencia",
    "IPJ", "Prioridad SEO", "Prioridad Ads",
]

# Patrón del aviso automático de cobertura, para poder separarlo de cualquier
# nota manual que alguien haya escrito en la misma celda de Observaciones.
PATRON_AVISO_COBERTURA = re.compile(
    r"Cobertura \d+% \(secreto estadístico parcial\) — verificar antes de segmentar\.?"
)


def leer_datos_previos(ws) -> dict:
    """Lee la hoja Ranking_CP tal como está antes de tocarla y devuelve,
    por Código Postal (normalizado a 5 dígitos), los valores de las
    columnas manuales y el texto libre de Observaciones."""
    previos = {}
    if ws.max_row < 2:
        return previos

    col_cp = COL_IDX["Código Postal"]
    col_obs = COL_IDX["Observaciones"]

    for r in range(2, ws.max_row + 1):
        cp_val = ws.cell(r, col_cp).value
        if cp_val is None or str(cp_val).strip() == "":
            continue
        cp = str(cp_val).strip().zfill(5)

        datos = {}
        for nombre in COLUMNAS_MANUALES:
            datos[nombre] = ws.cell(r, COL_IDX[nombre]).value

        obs_actual = ws.cell(r, col_obs).value or ""
        obs_manual = PATRON_AVISO_COBERTURA.sub("", obs_actual).strip(" -—")
        datos["Observaciones_manual"] = obs_manual

        previos[cp] = datos

    return previos


def main():
    if not INFORME.exists():
        raise FileNotFoundError(f"No se encontró {INFORME.name}.")
    if not PLANTILLA.exists():
        raise FileNotFoundError(f"No se encontró {PLANTILLA.name}.")

    df = pd.read_excel(INFORME, dtype={"Codigo_Postal": str, "Codigo_INE_Principal": str})

    wb = load_workbook(PLANTILLA)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"La plantilla no contiene la hoja '{HOJA}'.")
    ws = wb[HOJA]

    # 1. Guardar lo que ya había, ANTES de borrar nada.
    datos_previos = leer_datos_previos(ws)

    # 2. Ahora sí, limpiar el cuerpo de la hoja para reconstruirlo.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    cp_vistos = set()
    for i, fila in df.iterrows():
        r = i + 2
        cp = str(fila["Codigo_Postal"]).zfill(5)
        cp_vistos.add(cp)
        previo = datos_previos.get(cp, {})

        cobertura_baja = fila.get("Pct_Cobertura_Renta", 100) < 50
        aviso_auto = (
            f"Cobertura {fila.get('Pct_Cobertura_Renta', 0):.0f}% "
            f"(secreto estadístico parcial) — verificar antes de segmentar."
            if cobertura_baja else ""
        )
        obs_manual = previo.get("Observaciones_manual", "")
        observacion = " ".join(p for p in [aviso_auto, obs_manual] if p).strip()

        ws.cell(r, COL_IDX["Ranking"], i + 1)
        ws.cell(r, COL_IDX["Código Postal"], cp)
        ws.cell(r, COL_IDX["Municipio"], fila["Municipio"])
        ws.cell(r, COL_IDX["Codigo INE"], fila.get("Codigo_INE_Principal", ""))

        # Renta por hogar y evoluciones (siempre las recalcula el pipeline)
        ws.cell(r, COL_IDX["Renta Bruta media (€)"], fila.get("Renta_Bruta_Hogar_Media"))
        ws.cell(r, COL_IDX["Evolución bruta"], fila.get("Evolución_Bruta_%"))
        ws.cell(r, COL_IDX["Renta Neta media (€)"], fila.get("Renta_Neta_Hogar_Media"))
        ws.cell(r, COL_IDX["Evolucion Neta"], fila.get("Evolución_Neta_%"))
        ws.cell(r, COL_IDX["Urbanizaciones"], fila.get("Pct_Urbanizaciones"))

        # Columnas manuales: se restauran tal cual estaban para ese CP.
        for nombre in COLUMNAS_MANUALES:
            ws.cell(r, COL_IDX[nombre], previo.get(nombre))

        ws.cell(r, COL_IDX["Observaciones"], observacion)

    wb.save(PLANTILLA)

    def tiene_datos_manuales(previo: dict) -> bool:
        return any(previo.get(c) not in (None, "") for c in COLUMNAS_MANUALES) or bool(
            previo.get("Observaciones_manual")
        )

    cp_desaparecidos = set(datos_previos) - cp_vistos
    cp_con_datos_conservados = sum(
        1 for cp in cp_vistos if cp in datos_previos and tiene_datos_manuales(datos_previos[cp])
    )
    print(f"Plantilla '{PLANTILLA.name}' actualizada correctamente con {len(df)} CP.")
    print(f"  CP con columnas manuales conservadas: {cp_con_datos_conservados}")
    if cp_desaparecidos:
        print(f" AVISO: {len(cp_desaparecidos)} CP que tenían datos manuales ya no "
              f"aparecen en el informe nuevo (se han perdido sus columnas manuales): "
              f"{', '.join(sorted(cp_desaparecidos))}")


if __name__ == "__main__":
    main()